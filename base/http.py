# -*- coding: utf-8 -*-
# Author:Paul J
# Verson:3.0.0
from __future__ import unicode_literals
import requests
from RequestsLibrary import RequestsKeywords
from Automation.utils.util import logger, abspath, getPythonDir, ConfigIni
import json, os, csv, datetime, time, shutil, sys, xml.dom.minidom
from ssh import SSHClient

reload(sys)
sys.setdefaultencoding("utf-8")


class Jmeter(object):
    '''
    Jmeter相关关键字方法
    '''
    # 用于记录性能测试场景jmx文件的列表#
    __jmx_list = []
    # 用于记录性能测试场景jmx文件的列表(不带编号)#
    __jmx_list2 = []
    # 性能测试时，记录xml路径#
    __xml_path = ""
    # 性能测试时，记录聚合报告路径#
    __report_path = ""
    # 性能测试时，记录服务器资源csv路径#
    __perf_mon_path = ""
    # 性能测试时，记录csv文件路径#
    __csv_path = ""
    # 存放本次性能测试的测试报告文件名#
    __performance_test_report_filename = ""
    # 性能测试时，记录编号#
    __jmx_num = ""
    # 性能测试时，记录开始时间#
    __start_time = ""
    # 性能测试，记录结束时间#
    __end_time = ""

    def modifyParameter(self, jmx_name, **params):
        '''
        - 功能：修改jmeter脚本运行时参数，包括参数传递项：线程数、循环数、运行时间、地址-IPADDRESS变量，自动修改项：查看结果树开启xml、开启记录响应数据、查看结果树保存地址、聚合报告地址
        - 参数如下：
        - jmx_name：jmx脚本名称，不带扩展名，切一定要唯一
        - loop（可选）：循环数
        - thread（可选）：线程数
        - time（可选）：持续时间，单位秒
        - address（可选）：需要修改的url地址
        - RobotFramwork脚本写法示例如下：
        | modifyParameter | appService | loop=3 | thread=2 | time=2 | address=www.hh.com |
        '''
        jmx_path = abspath(__file__, "../resource/jmeter/" + jmx_name + ".jmx")
        # 普通接口测试#
        if self.__performance_test_report_filename == "":
            xml_path = abspath(__file__, "../resource/jmeter/output/" + jmx_name + ".xml")
            report_path = abspath(__file__, "../resource/jmeter/output/" + jmx_name + "_report.csv")
        # 性能测试#
        else:
            t = time.time()
            self.__jmx_num = "@@" + str(int(t))
            xml_path = abspath(__file__, "../resource/jmeter/output/" + jmx_name + self.__jmx_num + ".xml")
            report_path = abspath(__file__, "../resource/jmeter/output/" + jmx_name + self.__jmx_num + "_report.csv")
            perf_mon_path = abspath(__file__, "../resource/jmeter/output/" + jmx_name + self.__jmx_num + "_perf.csv")
            self.__xml_path = xml_path
            self.__report_path = report_path
            self.__perf_mon_path = perf_mon_path
        dom = xml.dom.minidom.parse(jmx_path)
        root = dom.documentElement
        # 找线程组#
        elem_ThreadGroups = root.getElementsByTagName('ThreadGroup')
        for item in elem_ThreadGroups:
            if item.getAttribute("enabled") == "true":
                elem_ThreadGroup = item
        # 找配置项#
        itemlist = elem_ThreadGroup.getElementsByTagName('stringProp')
        # 修改循环数#
        if 'loop' in params and params['loop'] != '':
            for item in itemlist:
                if item.getAttribute("name") == "LoopController.loops":
                    logger.info("修改循环数量：" + params['loop'])
                    item.childNodes[0].nodeValue = params['loop']
        # 修改线程数#
        if 'thread' in params and params['thread'] != '':
            for item in itemlist:
                if item.getAttribute("name") == "ThreadGroup.num_threads":
                    logger.info("修改线程数：" + params['thread'])
                    item.childNodes[0].nodeValue = params['thread']
        # 修改时间#
        if 'time' in params and params['time'] != '':
            for item in itemlist:
                if item.getAttribute("name") == "ThreadGroup.ramp_time":
                    logger.info("修改持续时间：" + params['time'] + "秒 ")
                    item.childNodes[0].nodeValue = params['time']
        # 改地址#
        if "address" in params and params['address'] != '':
            elem_elementProps = root.getElementsByTagName('elementProp')
            for element in elem_elementProps:
                if element.getAttribute("name") == "IPADDRESS":
                    address_elem = element
            try:
                pros = address_elem.getElementsByTagName('stringProp')
            except UnboundLocalError:
                raise ValueError("jmx脚本缺少以IPADDRESS命名的自定义变量")
            for item in pros:
                if item.getAttribute("name") == "Argument.value":
                    logger.info("修改接口地址：" + params['address'])
                    item.childNodes[0].nodeValue = params['address']
        # 获取所有监听器#
        listers = root.getElementsByTagName('ResultCollector')
        # 找最后一个聚合报告#
        juhebaogao = []
        for item in listers:
            if item.getAttribute("guiclass") == "StatVisualizer" and item.getAttribute("enabled") == "true":
                juhebaogao.append(item)
        # 最后一个才是要改的#
        if juhebaogao == []:
            raise ValueError("jmx脚本缺少有效的聚合报告器件，请检查")
        else:
            result_report = juhebaogao.pop()
        # 修改结果树保存路径#
        pro_filename = result_report.getElementsByTagName('stringProp')
        for item in pro_filename:
            try:
                logger.info("修改聚合报告保存路径：" + report_path)
                item.childNodes[0].nodeValue = report_path
            except:
                logger.info("创建TextNode节点路径：" + report_path)
                # 创建新的textnode追加到节点上！！！一定是dom上创建#
                newtext = dom.createTextNode(report_path);
                item.appendChild(newtext)
        # 找最后一个查看结果树#
        jieguoshu = []
        for item in listers:
            if item.getAttribute("guiclass") == "ViewResultsFullVisualizer" and item.getAttribute("enabled") == "true":
                jieguoshu.append(item)
        # 最后一个才是要改的#
        if jieguoshu == []:
            raise ValueError("jmx脚本缺少有效的查看结果树器件，请检查")
        else:
            result_tree = jieguoshu.pop()
        # 修改结果树保存路径#
        pro_filename = result_tree.getElementsByTagName('stringProp')
        for item in pro_filename:
            try:
                logger.info("修改结果树保存路径：" + xml_path)
                item.childNodes[0].nodeValue = xml_path
            except:
                logger.info("创建TextNode节点路径：" + xml_path)
                # 创建新的textnode追加到节点上！！！一定是dom上创建#
                newtext = dom.createTextNode(xml_path);
                item.appendChild(newtext);
        # 修改开启xml#
        conf_xml = result_tree.getElementsByTagName('xml')
        for item in conf_xml:
            logger.info("开启xml")
            item.childNodes[0].nodeValue = "true"
        # 修改保存响应结果#
        conf_response = result_tree.getElementsByTagName('responseData')
        logger.info("保存响应结果")
        for item in conf_response:
            item.childNodes[0].nodeValue = "true"
        # 性能测试时，修改资源监视器#
        if self.__performance_test_report_filename != "":
            # 找资源监视器#
            profs = root.getElementsByTagName('kg.apc.jmeter.perfmon.PerfMonCollector')
            for item in profs:
                if item.getAttribute("enabled") == "true":
                    profs = item
            if profs == [] or profs is None:
                logger.info(jmx_name + "场景无服务器性能监视器")
            else:
                # 1.保存文件的找配置项#
                itemlist = profs.getElementsByTagName('stringProp')
                for item in itemlist:
                    if item.getAttribute("name") == "filename":
                        try:
                            logger.info("修改资源监视器保存路径：" + perf_mon_path)
                            item.childNodes[0].nodeValue = perf_mon_path
                        except:
                            logger.info("创建TextNode节点路径：" + perf_mon_path)
                            # 创建新的textnode追加到节点上！！！一定是dom上创建#
                            newtext = dom.createTextNode(perf_mon_path);
                            item.appendChild(newtext)
                # 2.找ip修改的配置项#
                itemlist = profs.getElementsByTagName('collectionProp')
                for item in itemlist:
                    if item.getAttribute("name") == "metricConnections":
                        sub_item = item.getElementsByTagName('collectionProp')
                        for i in sub_item:
                            i_tmps = i.getElementsByTagName('stringProp')
                            if i_tmps[0].childNodes[0].nodeValue != "${IPADDRESS}":
                                logger.info("需要修改资源监视器监视地址：" + params['address'])
                                i_tmps[0].childNodes[0].nodeValue = params['address']
        f = open(jmx_path, 'w')
        dom.writexml(f, encoding='utf-8')

    def runJmeterScript(self, jmeter_name, param=None):
        '''
        - 功能：执行Jmeter脚本(版本3.2)
        - 参数：
        | jmeter_name：jmx名，不加后缀名  | csv_name：生成csv文件 | param：需要附加的命令行参数 |
        '''
        jmx_path = abspath(__file__, "../resource/jmeter/" + jmeter_name + ".jmx")
        logger.info("Jmeter脚本路径：" + jmx_path)
        # 普通接口测试#
        if self.__performance_test_report_filename == "":
            csv_path = abspath(__file__, "../resource/jmeter/output/" + jmeter_name + ".csv")
            xml_path = abspath(__file__, "../resource/jmeter/output/" + jmeter_name + ".xml")
            report_path = abspath(__file__, "../resource/jmeter/output/" + jmeter_name + "_report.csv")
        # 性能测试#
        else:
            csv_path = abspath(__file__, "../resource/jmeter/output/" + jmeter_name + self.__jmx_num + ".csv")
            xml_path = self.__xml_path
            report_path = self.__report_path
        logger.info("Jmeter结果文件保存路径：" + csv_path)
        if os.path.exists(csv_path):
            logger.info("CSV历史记录存在，需清除")
            os.remove(csv_path)
        if os.path.exists(xml_path):
            logger.info("xml历史记录存在，需清除")
            os.remove(xml_path)
        if os.path.exists(report_path):
            logger.info("聚合报告历史记录存在，需清除")
            os.remove(report_path)
        if param is None:
            cmd = "jmeter -n -t " + jmx_path + " -l " + csv_path
        else:
            cmd = "jmeter -n -t " + jmx_path + " -l " + csv_path + " " + param
        logger.info("Jmeter执行命令：" + cmd)
        logger.info("Jmter正在执行%s脚本，请耐心等待……" % (jmeter_name,))
        result = os.popen(cmd.decode("utf8").encode("gbk"))
        logger.info("Jmeter执行结果：")
        for line in result.readlines():
            try:
                logger.info(line.strip("\n"))
            except:
                logger.info("编码问题忽略：Created the tree successfully using：" + jmx_path)
        # 仅仅在性能测试时候#
        if self.__performance_test_report_filename != "":
            self.__jmx_list.append(jmeter_name + self.__jmx_num)
            self.__jmx_list2.append(jmeter_name)

    def checkJmeterResult(self, jmeter_name):
        '''
        - 功能：通过解析xml日志文件，分析jmeter执行结果，包括http和java请求
        - 参数：
        | jmeter_name：日志文件名  |
        '''
        xml_path = abspath(__file__, "../resource/jmeter/output/" + jmeter_name + ".xml")
        logger.info("Jmeter结果文件保存路径(xml)：" + xml_path)
        if not os.path.exists(xml_path):
            logger.info("xml结果文件不存在")
            return False
        dom = xml.dom.minidom.parse(xml_path)
        root = dom.documentElement
        results = []  # 存放所有的结果集
        false_flag = 0  # 存放单个接口失败标志
        return_flag = 0  # 统计多少个接口失败
        # 找线程组#
        itf_logs = root.getElementsByTagName('httpSample')
        if itf_logs == []:
            itf_logs = root.getElementsByTagName('sample')
        for one_log in itf_logs:
            result = []
            false_flag = 0
            # 保存测试结果#
            test_result = one_log.getAttribute("s")
            if test_result == "true":
                result.append("通过")
                logger.info("-------------------------")
            else:
                result.append("失败")
                false_flag = 1
                return_flag = return_flag + 1
                logger.warn("#########################")
                logger.warn("[存在失败接口信息如下]：")
            # 保存时间#
            timeStamp = int(one_log.getAttribute("ts")) / 1000
            datetime_struct = str(datetime.datetime.fromtimestamp(timeStamp))
            result.append(datetime_struct)
            if false_flag == 1:
                logger.warn("调用时间：" + datetime_struct)
            else:
                logger.info("调用时间：" + datetime_struct)
            # 保存场景名#
            changjing = one_log.getAttribute("tn")
            result.append(changjing)
            if false_flag == 1:
                logger.warn("场景名称：" + changjing)
            else:
                logger.info("场景名称：" + changjing)
            # 保存接口名#
            itf_name = one_log.getAttribute("lb")
            result.append(itf_name)
            if false_flag == 1:
                logger.warn("接口名称：" + itf_name)
            else:
                logger.info("接口名称：" + itf_name)
            # 保存响应结果码#
            result_code = one_log.getAttribute("rc")
            if result_code == "":
                result_code = "None"
            result.append(result_code)
            if false_flag == 1:
                logger.warn("响应结果码：" + result_code)
            else:
                logger.info("响应结果码：" + result_code)
            # 保存响应结果错误信息#
            fail_message = one_log.getAttribute("rm")
            if fail_message == "":
                fail_message = "None"
            result.append(fail_message)
            if false_flag == 1:
                logger.warn("响应结果错误信息：" + fail_message)
            else:
                logger.info("响应结果错误信息：" + fail_message)
            # 获取响应数据#
            response_data = one_log.getElementsByTagName('responseData')
            for item in response_data:
                try:
                    out = item.childNodes[0].nodeValue
                    response_data = self.__jsonPrettyOutput(out)
                except Exception, e:
                    logger.warn(e)
                    response_data = out
                result.append(response_data)
                if false_flag == 1:
                    logger.warn("响应报文：" + response_data)
                else:
                    logger.info("响应报文：" + response_data)
            results.append(result)
        if return_flag > 0:
            logger.info("Jmeter脚本执行存在%d个错误接口" % (return_flag))
            return False
        else:
            logger.info("Jmeter脚本执行正确")
            return True

    def initPerformanceTest(self, filename="性能测试报告"):
        '''
        - 功能：初始化性能测试，创建空的性能测试报告（与createPerformanceTestReport配对）
        - 参数：
        | filename：性能测试报告文件名，可选  |
        '''
        self.__performance_test_report_filename = ""
        report_tmp = abspath(__file__, "../resource/jmeter/report_template/性能测试结果模板.xlsm")
        now_time = time.strftime("%Y%m%d%H%M%S", time.localtime(time.time()))
        report_path = abspath(__file__, "../resource/jmeter/output/" + filename + "_" + now_time + ".xlsm")
        self.__performance_test_report_filename = filename + "_" + now_time + ".xlsm"
        logger.info("初始化本次性能测试：" + now_time)
        shutil.copy(report_tmp, report_path)
        try:
            import xlwt
        except:
            logger.info("无xlwt模块，需要联网下载")
            os.system(getPythonDir() + "/Scripts/pip install xlwt -i https://pypi.douban.com/simple/")
        try:
            import xlrd
        except:
            logger.info("无xlrd模块，需要联网下载")
            os.system(getPythonDir() + "/Scripts/pip install xlrd -i https://pypi.douban.com/simple/")
        try:
            from openpyxl import load_workbook
        except:
            logger.info("无openpyxl模块，需要联网下载")
            os.system(getPythonDir() + "/Scripts/pip install openpyxl -i https://pypi.douban.com/simple/")
        # 清理历史过程数据#
        logger.info("清理历史过程数据")
        dir = abspath(__file__, "../resource/jmeter/output/")
        cmd1 = "del /s /f /q " + dir + "\*@@*.*"
        os.system(cmd1.decode("utf8").encode("gbk"))
        self.__start_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        logger.info("性能测试开始时间：" + self.__start_time)

    def __getPerfInfo(self, jmeter_name):
        '''
        - 功能：分析性能资源数据
        - 参数：
        | jmeter_name：jmx名  |
        
        '''
        csv_perf_path = abspath(__file__, "../resource/jmeter/output/" + jmeter_name + "_perf.csv")
        csvFile = open(csv_perf_path, "r")
        reader = csv.reader(csvFile)
        csv_list = []
        # 把每一行加到列表中#
        for line in reader:
            # 忽略第一行标题#
            if reader.line_num == 1:
                continue
            csv_list.append(line)
        csvFile.close()
        # 先取内存的最大#
        M_max = 0
        for line in csv_list:

            if "Memory" in line[2]:
                if int(line[1]) > M_max:
                    M_max = int(line[1])
        CPU_max = 0
        for line in csv_list:
            if "CPU" in line[2]:
                if int(line[1]) > CPU_max:
                    CPU_max = int(line[1])
        return (
            str(round(float(CPU_max / 1000), 2)),
            str(round(float(M_max / 1000), 2))
        )

    def __analysisCsvReport(self, jmeter_name):
        '''
        - 功能：分析聚合报告
        - 参数：
        | jmeter_name：聚合报告名  |
        '''
        csv_report_path = abspath(__file__, "../resource/jmeter/output/" + jmeter_name + "_report.csv")
        csvFile = open(csv_report_path, "r")
        reader = csv.reader(csvFile)
        csvList = []
        try:
            perf_info = self.__getPerfInfo(jmeter_name)
        except IOError:
            perf_info = ("0.0", "0.0")
        # 把每一行加到列表中#
        for line in reader:
            # 忽略第一行标题#
            if reader.line_num == 1:
                continue
            csvList.append(line)
        csvFile.close()
        count = len(csvList)
        reproList = []
        num = 0
        errNum = 0
        # 计算请求总数、错误总数#
        for row in csvList:
            data = long(row[10])
            failure = row[7]
            if failure == "false":
                errNum = errNum + 1
            num = num + data
        # 请求总数
        logger.info("[脚本名称]：" + jmeter_name)
        reproList.append(str(count))
        logger.info("请求总数：" + str(count))
        # 计算CPU#
        reproList.append(perf_info[0])
        logger.info("CPU：" + perf_info[0])
        # 计算内存#
        reproList.append(perf_info[1])
        logger.info("内存：" + perf_info[1])
        # 计算平均响应
        reproList.append(str(round(long(num / count), 2)))
        logger.info("平均响应时间：" + str(round(long(num / count), 2)) + "ms")
        # 计算吞吐量
        aLong = long(csvList[0][0])
        bLong = long(csvList[len(csvList) - 1][0])
        if bLong != aLong:
            thought = str(round(count / ((bLong - aLong) / 1000.0), 2))
            reproList.append(thought)
            logger.info("吞吐量：" + thought + "/s")
        else:
            reproList.append("Null")
            logger.info("吞吐量：只有一次请求，无法计算吞吐量")
        # 计算错误率
        reproList.append(str(round(long(errNum / count), 2)))
        logger.info("计算错误率：" + str(round(long(errNum / count), 2)))
        logger.info("----------------------------------------")
        return reproList

    def __prettyExcelTemplate(self, file_path):
        '''
        - 功能：美化性能测试报告
        - 参数：
        | file_path：性能测试 报告路径 |
        '''
        logger.info("要美化性能测试报告")
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side, Font
        # 读取excel保留VBA宏
        wb2 = load_workbook(file_path, keep_vba=True)
        # 激活默认第一页sheet
        ws = wb2.active
        # 设置线条格式
        border = Border(left=Side(style='thin', color='FF000000'),
                        right=Side(style='thin', color='FF000000'),
                        top=Side(style='thin', color='FF000000'),
                        bottom=Side(style='thin', color='FF000000')
                        )
        # 给A1到H10区域进行加边框
        cell_range = ws['A1:H10']
        for row in cell_range:
            for cell in row:
                cell.border = border
        # 保存文件
        wb2.save(file_path)

    def createPerformanceTestReport(self):
        '''
        - 功能：创建性能测试结果报告（与initPerformanceTest配对）
        '''
        import xlwt
        import xlrd
        from openpyxl import load_workbook
        from openpyxl.chart import BarChart, Series, Reference, BarChart3D
        from openpyxl.styles import Border, Side, Font, Alignment
        self.__end_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        logger.info("性能测试结束时间：" + self.__end_time)
        # 只有一个场景，warn提示，不能生成折线图的#
        for item in self.__jmx_list2:
            num = self.__jmx_list2.count(item)
            if num == 1:
                logger.warn(item.decode("utf-8") + "脚本只有一个场景，故无法生成趋势折现图。")
        if self.__performance_test_report_filename == "":
            raise ValueError("未执行过initPerformanceTest方法")
        if self.__jmx_list == "":
            raise ValueError("jmx脚本执行异常，请检查")
        report_path = abspath(__file__, "../resource/jmeter/output/" + self.__performance_test_report_filename)
        index = 1
        # 合并单元格时候的标记#
        now_jmx_flag = ""
        # 打开excel#
        wb2 = load_workbook(report_path, keep_vba=True)
        ws = wb2.active
        for one_jmx in self.__jmx_list:
            # 计算各个jmx的聚合报告中的信息#
            reslut_list = self.__analysisCsvReport(one_jmx)
            # 如果是合并单元格，要合并的位移量#
            num = self.__jmx_list2.count(one_jmx.split("@@")[0])
            # 第一列，作缩进#
            if now_jmx_flag.split("@@")[0] != one_jmx.split("@@")[0]:
                # 合并单元格#
                ws.cell(row=11 + index, column=1, value=one_jmx.split("@@")[0].decode("utf-8"))
                # 水平、垂直居中#
                ws.cell(row=11 + index, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=11 + index, end_row=11 + index + num - 1, start_column=1, end_column=1)
                # 做标记#
                now_jmx_flag = one_jmx.split("@@")[0]
            ws.cell(row=11 + index, column=2, value=int(reslut_list[0]))
            ws.cell(row=11 + index, column=3, value=float(reslut_list[1]))
            ws.cell(row=11 + index, column=4, value=float(reslut_list[2]))
            ws.cell(row=11 + index, column=5, value=float(reslut_list[3]))
            try:
                ws.cell(row=11 + index, column=6, value=float(reslut_list[4]))
            except:
                ws.cell(row=11 + index, column=6, value=reslut_list[4])
            ws.cell(row=11 + index, column=7, value=float(reslut_list[5]))
            index = index + 1
        ws["F2"].value = self.__start_time + " —— " + self.__end_time
        # 保存结果#
        wb2.save(report_path)
        logger.info("性能测试报告：%s创建完成" % (self.__performance_test_report_filename,))
        self.__prettyExcelTemplate(report_path)


class locust(object):
    '''
    Locust相关关键字方法
    '''

    def __init__(self):
        '''
        -- 初始化：导入locust包
        '''
        try:
            import locust
        except:
            logger.write_info_log("##### No locust package,need to download. #####")
            os.system(getPythonDir() + "/Scripts/pip install locustio")
            os.system(getPythonDir() + "/Scripts/pip install locust")

    def createConfigure(self, url, uri, **kwargs):
        '''
        - 功能：将参数写进locust执行配置文件
        - 参数如下：
        - url：接口url（必填）
        - uri：接口uri（必填）
        - header：接口请求header信息，默认{"Content-Type":"application/json"}
        - type：接口请求方法（POST/GET），默认POST
        - data：POST请求时，请求body的json串，默认为None
        - port：web模式下，locust访问端口，默认8089
        - no_web：是否启用Web模式，默认False，即命令行模式
        - num：并发用户数，默认20（no_web模式）
        - rate：每秒启动用户数，默认1（no_web模式）
        - run_time：脚本运行时间，默认10s（no_web模式）
        - min_wait：用户执行任务之间等待时间的下界，单位：毫秒，默认2
        - max_wait：用户执行任务之间等待时间的上界，单位：毫秒，默认5
        - checkpoint：返回json串校验，以key=>value形式传递，默认为None
        - RobotFramwork脚本写法示例如下：
        | createConfigure | http://10.0.0.205:8088 | transform | data={ "group":"user-center-service","method":"authUserByMobile","params":["3UDO5RAHPMBQEODP7ILJV3LLZY3ZK","18700000006","123456"],"service":"cn.com.flaginfo.platform.ucenter.api.AuthService","timeout":3000, "version":"6.1.0"} | checkpoint=message=>用户验证成功！ |
        '''
        param_dict = {}
        # url#
        param_dict["url"] = url
        # uri#
        if uri[0:1] != "/":
            uri = "/" + uri
        param_dict["uri"] = uri
        # header(json格式)#
        if "header" in kwargs and kwargs["header"] != "":
            param_dict["header"] = kwargs["header"]
        else:
            logger.write_info_log("##### Use default request header-{\"Content-Type\":\"application/json\"}. #####")
            param_dict["header"] = '{"Content-Type":"application/json"}'
        # POST/GET#
        if "type" in kwargs and kwargs["type"] != "":
            param_dict["type"] = kwargs["type"]
        else:
            logger.write_info_log("##### Use default request type-POST. #####")
            param_dict["type"] = 'POST'
        # data#
        if "data" in kwargs and kwargs["data"] != "":
            param_dict["data"] = kwargs["data"]
        else:
            param_dict["data"] = 'None'
        # 端口#
        if "port" in kwargs and kwargs["port"] != "":
            param_dict["port"] = kwargs["port"]
        else:
            logger.write_info_log("##### Use default port-8089. #####")
            param_dict["port"] = "8089"
        # 启动模式#
        if "no_web" in kwargs and kwargs["no_web"] != "":
            param_dict["no_web"] = kwargs["no_web"]
        else:
            logger.write_info_log("##### Use default mode-no_web. #####")
            param_dict["no_web"] = "True"
        # 并发用户数#
        if "num" in kwargs and kwargs["num"] != "":
            param_dict["num"] = kwargs["num"]
        else:
            if param_dict["no_web"] == "True":
                logger.write_info_log("##### Use default thread num-20. #####")
                param_dict["num"] = "20"
            else:
                logger.write_info_log("##### Web mode no need to config thread num. #####")
                param_dict["num"] = "1"
        # 每秒启动用户数#
        if "rate" in kwargs and kwargs["rate"] != "":
            param_dict["rate"] = kwargs["rate"]
        else:
            if param_dict["no_web"] == "True":
                logger.write_info_log("##### Use default rate-1. #####")
                param_dict["rate"] = "1"
            else:
                logger.write_info_log("##### Web mode no need to config rate. #####")
                param_dict["rate"] = "1"
        # 运行时间#
        if "run_time" in kwargs and kwargs["run_time"] != "":
            param_dict["run_time"] = kwargs["run_time"]
        else:
            if param_dict["no_web"] == "True":
                logger.write_info_log("##### Use default run_time-10s. #####")
                param_dict["run_time"] = "10s"
            else:
                logger.write_info_log("##### Web mode no need to config rate. #####")
                param_dict["run_time"] = "1"
        # 用户执行任务之间等待时间的下界，单位：毫秒#
        if "min_wait" in kwargs and kwargs["min_wait"] != "":
            param_dict["min_wait"] = kwargs["min_wait"]
        else:
            logger.write_info_log("##### Use default min_wait-5ms. #####")
            param_dict["min_wait"] = "2"
        # 用户执行任务之间等待时间的上界，单位：毫秒#
        if "max_wait" in kwargs and kwargs["max_wait"] != "":
            param_dict["max_wait"] = kwargs["max_wait"]
        else:
            logger.write_info_log("##### Use default max_wait-10ms. #####")
            param_dict["max_wait"] = "5"
        # 检查点校验#
        if "checkpoint" in kwargs and kwargs["checkpoint"] != "":
            param_dict["checkpoint"] = kwargs["checkpoint"]
        else:
            logger.write_info_log("##### No need to check json response content. #####")
            param_dict["checkpoint"] = "None"
        path = abspath(__file__, '../resource/locust/config.ini')
        # 创建ini配置文件，覆盖写#
        exec_file = open(path, 'w')
        exec_file.close()
        # 创建ini的title#
        cfg = ConfigIni(path, "config")
        cfg.add_ini("config")
        # 将执行配置，写入ini#
        for key in param_dict:
            cfg.set_ini("config", key, param_dict[key])

    def execLocust(self, script_file=None):
        '''
        - 功能：执行locust脚本
        - 参数如下：
        - script_file：Locust脚本所在位置，默认为None，自动读取resource/locust路径下的locustScript.py脚本
        - RobotFramwork脚本写法示例如下：
        | execLocust | d:\locust_script.py |
        '''
        path = abspath(__file__, '../resource/locust/config.ini')
        cfg = ConfigIni(path, "config")
        if cfg.get_ini("no_web") == "True":
            no_web = "--no-web"
            rate = "-r %s " % cfg.get_ini("rate")
            run_time = "-t %s " % cfg.get_ini("run_time")
            num = "-c %s " % cfg.get_ini("num")
        # 如果web模式，以下参数就不需要设置#
        else:
            no_web = ""
            rate = ""
            run_time = ""
            num = ""
            logger.write_info_log(
                "##### Locust work in web mode, you can access to http://localhost:%s to use. #####" % cfg.get_ini(
                    "port"))
        log_file = abspath(__file__, '../resource/locust/result')
        # 如果无特殊指定，locust脚本指向默认脚本#
        if script_file is None:
            script_file = abspath(__file__, '../resource/locust/locustScript.py')
        cmd = getPythonDir() + "/Scripts/locust -P %s -f %s %s %s %s %s --csv=%s" % (
            cfg.get_ini("port"), script_file, no_web, num, rate, run_time, log_file)
        logger.write_info_log("##### cmd #####")
        logger.write_info_log(cmd)
        if cfg.get_ini("no_web") == "True":
            os.system(cmd)
        else:
            # robot情况下，进入web模式，以子进程模式启动locust#
            import subprocess
            subprocess.Popen(cmd, shell=True)


class http(Jmeter, SSHClient, RequestsKeywords, locust):
    '''
    基本的http操作（可作为Robot关键字使用）
    '''
    # 存放校验json时的临时变量#
    __result_flag_itf = {}
    __result_flag_itf_desc = {}
    __check_type = "asc"
    # 存放校验字典包含时的临时变量#
    __result_contain_flag = []
    __result_contain_flag_desc = []

    def __init__(self):
        super(http, self).__init__()

    def HttpPost(self, url, uri, **kwargs):
        '''
        - 功能：发post请求
        - 参数如下：
        - url：接口url
        - uri：对应接口api
        - alias(可选)保持session会话的标识,默认为post_api
        - session_headers(可选):一个会话中公共请求头
        - session_cookies(可选)：一个会话中公共cookie请求头,格式类型为cookiejar，一般通过上次请求响应cookie获取
        - headers(可选): 发送POST请求的请求头，具体某一个请求，一个会话中可以有多个请求
        - data（可选）：请求体，格式为json
        - file_data(可选)：文件以请求体形式发送 ，格式为文件路径
        - files（可选）: 上传附件。格式为json，key为文件名，value为文件路径
        - allow_redirects(可选)：是否选择自动跳转。TRUE允许跳转，FALSE不允许跳转
        - context(可选):上下文联系参数,当需要进行上下文请求时选用该参数，与创建session中alias参数值保持一致
        - RobotFramwork脚本写法示例如下：
        | HttpPost | http://6.test.ums86.com/ | api/post.do | alias=post_api | session_cookies=${before_cookie} | session_headers={"Content-Type":"application/json"} | headers={"charset":"utf-8"} | data={"name": "liwei"} | file_data="C:/a.jpg" | files={"file1":"C:/a.txt"} | context="上文API" | allow_redirects=${false} |
        '''
        logger.info("##### Post [Url] %s [Uri] %s #####" % (url, uri))
        for key in kwargs:
            logger.info("##### Param-%s: %s #####" % (key, kwargs[key]))
        session_headers = {}
        session_cookies = None
        post_headers = None
        post_data = None
        post_files = {}
        post_allow_redirects = None
        alias = 'post_api'
        if 'alias' in kwargs and kwargs['alias'] != '':
            alias = kwargs['alias']
        if 'session_headers' in kwargs and kwargs['session_headers'] != '':
            session_headers = self.jsonToDict(kwargs['session_headers'])
        if 'session_cookies' in kwargs and kwargs['session_cookies'] != '':
            session_cookies = kwargs['session_cookies']
        if 'headers' in kwargs and kwargs['headers'] != '':
            post_headers = json.loads(kwargs['headers'])
        if 'data' in kwargs and kwargs['data'] != '':
            post_data = kwargs['data']
            # 响应结果格式化输出#
            try:
                pretty_output = self.__jsonPrettyOutput(post_data)
                logger.info("##### Request message:\n" + pretty_output)
            except Exception, e:
                logger.info("##### Request message:\n" + post_data)
                # logger.warn(e)
                # logger.warn("##### Request parameter is invalid json format:\n" + post_data)
        if 'file_data' in kwargs and kwargs['file_data'] != '':
            post_data = open(kwargs['file_data'], 'rb')
        if 'files' in kwargs and kwargs['files'] != '':
            files_dict = json.loads(kwargs['files'])
            for key in files_dict:
                post_files[key] = open(files_dict[key], 'rb')
        if 'allow_redirects' in kwargs and kwargs['allow_redirects'] != '':
            post_allow_redirects = kwargs['allow_redirects']
        if 'context' not in kwargs:
            self.create_session(alias, url, session_headers, cookies=session_cookies)
        else:
            alias = kwargs['context']
        requests.packages.urllib3.disable_warnings()
        itf_result = self.post_request(alias,
                                       uri,
                                       data=post_data,
                                       headers=post_headers,
                                       files=post_files,
                                       allow_redirects=post_allow_redirects
                                       )
        if 'context' not in kwargs:
            self.delete_all_sessions()
        return itf_result

    def HttpGet(self, url, uri, **kwargs):
        '''
        - 功能：发get请求
        - 参数如下：
        - url：接口url
        - uri：对应接口uri
        - alias(可选)保持session会话的标识,默认为get_api
        - session_headers(可选):一个会话中公共请求头，格式为json
        - session_cookies(可选)：一个会话中公共cookie请求头,格式为json
        - headers(可选): 发送GET请求的请求头，具体某一个请求，一个会话中可以有多个请求
        - params（可选）：请求参数，格式为json
        - json(可选)：以json方式发送请求体 ，格式为json
        - allow_redirects(可选)：是否选择自动跳转。TRUE允许跳转，FALSE不允许跳转
        - context(可选):上下文联系参数,当需要进行上下文请求时选用该参数，与创建session中alias参数值保持一致
        - RobotFramwork脚本写法示例如下：
        | HttpGet | http://6.test.ums86.com/ | api/getlist.do | alias=get_api | session_cookies=${before_cookie} | session_headers={"Content-Type":"application/x-www-form-urlencoded} | headers={"charset":"utf-8"} | params={"name": "liwei","password":"123"} | json={"data":"123"} | context="上文API" | allow_redirects=${false} |
        '''
        logger.info("##### Get [Url] %s [Uri] %s #####" % (url, uri))
        for key in kwargs:
            logger.info("##### Param-%s: %s #####" % (key, kwargs[key]))
        session_headers = {}
        session_cookies = None
        get_headers = {}
        get_params = None
        get_json = None
        get_allow_redirects = None
        alias = "get_api"
        if 'alias' in kwargs and kwargs['alias'] != '':
            alias = kwargs['alias']
        if 'session_headers' in kwargs and kwargs['session_headers'] != '':
            session_headers = json.loads(kwargs['session_headers'])
        if 'session_cookies' in kwargs and kwargs['session_cookies'] != '':
            session_cookies = kwargs['session_cookies']
        # 解析headers#
        if 'headers' in kwargs and kwargs["headers"] != "":
            get_headers = json.loads(kwargs["headers"])
        # 解析params#
        if "params" in kwargs and kwargs["params"] != "":
            get_params = json.loads(kwargs["params"])
        if "json" in kwargs and kwargs['json'] != "":
            get_json = json.loads(kwargs["json"])
        if "allow_redirects" in kwargs and kwargs['allow_redirects'] != "":
            get_allow_redirects = kwargs['allow_redirects']
        if 'context' not in kwargs:
            self.create_session(alias, url, session_headers, cookies=session_cookies)
        else:
            alias = kwargs['context']
        requests.packages.urllib3.disable_warnings()
        # 调get接口，保存结果#
        itf_result = self.get_request(alias, uri, headers=get_headers, json=get_json, params=get_params,
                                      allow_redirects=get_allow_redirects)
        if 'context' not in kwargs:
            self.delete_all_sessions()
        # 返回响应结果#
        return itf_result

    def HttpPut(self, url, uri, **kwargs):
        '''
        - 功能：发Put请求
        - 参数如下：
        - url：接口url
        - uri：对应接口api
        - alias(可选)保持session会话的标识,默认为post_api
        - session_headers(可选):一个会话中公共请求头
        - session_cookies(可选)：一个会话中公共cookie请求头,格式类型为cookiejar，一般通过上次请求响应cookie获取
        - headers(可选): 发送POST请求的请求头，具体某一个请求，一个会话中可以有多个请求
        - data（可选）：请求体，格式为json
        - file_data(可选)：文件以请求体形式发送 ，格式为文件路径
        - files（可选）: 上传附件。格式为json，key为文件名，value为文件路径
        - allow_redirects(可选)：是否选择自动跳转。TRUE允许跳转，FALSE不允许跳转
        - context(可选):上下文联系参数,当需要进行上下文请求时选用该参数，与创建session中alias参数值保持一致
        - RobotFramwork脚本写法示例如下：
        | HttpPost | http://6.test.ums86.com/ | api/post.do | alias=post_api | session_cookies=${before_cookie} | session_headers={"Content-Type":"application/json"} | headers={"charset":"utf-8"} | data={"name": "liwei"} | file_data="C:/a.jpg" | files={"file1":"C:/a.txt"} | context="上文API" | allow_redirects=${false} |
        '''
        logger.info("##### Put [Url] %s [Uri] %s #####" % (url, uri))
        for key in kwargs:
            logger.info("##### Param-%s: %s #####" % (key, kwargs[key]))
        session_headers = {}
        session_cookies = None
        put_headers = None
        put_data = None
        put_files = {}
        put_allow_redirects = None
        alias = 'put_api'
        if 'alias' in kwargs and kwargs['alias'] != '':
            alias = kwargs['alias']
        if 'session_headers' in kwargs and kwargs['session_headers'] != '':
            session_headers = self.jsonToDict(kwargs['session_headers'])
        if 'session_cookies' in kwargs and kwargs['session_cookies'] != '':
            session_cookies = kwargs['session_cookies']
        if 'headers' in kwargs and kwargs['headers'] != '':
            put_headers = json.loads(kwargs['headers'])
        if 'data' in kwargs and kwargs['data'] != '':
            put_data = kwargs['data']
            # 响应结果格式化输出#
            try:
                pretty_output = self.__jsonPrettyOutput(put_data)
                logger.info("##### Request message:\n" + pretty_output)
            except Exception, e:
                logger.info("##### Request message:\n" + put_data)
                # logger.warn(e)
                # logger.warn("##### Request parameter is invalid json format:\n" + post_data)
        if 'file_data' in kwargs and kwargs['file_data'] != '':
            put_data = open(kwargs['file_data'], 'rb')
        if 'files' in kwargs and kwargs['files'] != '':
            files_dict = json.loads(kwargs['files'])
            for key in files_dict:
                put_files[key] = open(files_dict[key], 'rb')
        if 'allow_redirects' in kwargs and kwargs['allow_redirects'] != '':
            put_allow_redirects = kwargs['allow_redirects']
        if 'context' not in kwargs:
            self.create_session(alias, url, session_headers, cookies=session_cookies)
        else:
            alias = kwargs['context']
        requests.packages.urllib3.disable_warnings()
        itf_result = self.put_request(alias,
                                      uri,
                                      data=put_data,
                                      headers=put_headers,
                                      files=put_files,
                                      allow_redirects=put_allow_redirects
                                      )
        if 'context' not in kwargs:
            self.delete_all_sessions()
        return itf_result

    def HttpDelete(self, url, uri, **kwargs):
        '''
        - 功能：发Delete请求
        - 参数如下：
        - url：接口url
        - uri：对应接口uri
        - alias(可选)保持session会话的标识,默认为get_api
        - session_headers(可选):一个会话中公共请求头，格式为json
        - session_cookies(可选)：一个会话中公共cookie请求头,格式为json
        - headers(可选): 发送GET请求的请求头，具体某一个请求，一个会话中可以有多个请求
        - params（可选）：请求参数，格式为json
        - json(可选)：以json方式发送请求体 ，格式为json
        - allow_redirects(可选)：是否选择自动跳转。TRUE允许跳转，FALSE不允许跳转
        - context(可选):上下文联系参数,当需要进行上下文请求时选用该参数，与创建session中alias参数值保持一致
        - RobotFramwork脚本写法示例如下：
        | HttpDelete | http://6.test.ums86.com/ | api/getlist.do | alias=get_api | session_cookies=${before_cookie} | session_headers={"Content-Type":"application/x-www-form-urlencoded} | headers={"charset":"utf-8"} | params={"name": "liwei","password":"123"} | json={"data":"123"} | context="上文API" | allow_redirects=${false} |
        '''
        logger.info("##### Delete [Url] %s [Uri] %s #####" % (url, uri))
        for key in kwargs:
            logger.info("##### Param-%s: %s #####" % (key, kwargs[key]))
        session_headers = {}
        session_cookies = None
        delete_params = None
        delete_data = None
        delete_headers = None
        delete_allow_redirects = None
        alias = "Delete_api"
        if 'alias' in kwargs and kwargs['alias'] != '':
            alias = kwargs['alias']
        if 'session_headers' in kwargs and kwargs['session_headers'] != '':
            session_headers = json.loads(kwargs['session_headers'])
        if 'session_cookies' in kwargs and kwargs['session_cookies'] != '':
            session_cookies = kwargs['session_cookies']
        # 解析headers#
        if 'headers' in kwargs and kwargs["headers"] != "":
            delete_headers = json.loads(kwargs["headers"])
        # 解析params#
        if "params" in kwargs and kwargs["params"] != "":
            delete_params = json.loads(kwargs["params"])
        if 'data' in kwargs and kwargs['data'] != '':
            delete_data = kwargs['data']
            # 响应结果格式化输出#
            try:
                pretty_output = self.__jsonPrettyOutput(delete_data)
                logger.info("##### Request message:\n" + pretty_output)
            except Exception, e:
                logger.info("##### Request message:\n" + delete_data)
                # logger.warn(e)
                # logger.warn("##### Request parameter is invalid json format:\n" + post_data)
        else:
            delete_data = ()
        if "allow_redirects" in kwargs and kwargs['allow_redirects'] != "":
            delete_allow_redirects = kwargs['allow_redirects']
        if 'context' not in kwargs:
            self.create_session(alias, url, session_headers, cookies=session_cookies)
        else:
            alias = kwargs['context']
        requests.packages.urllib3.disable_warnings()
        # 调delete接口，保存结果#
        itf_result = self.delete_request(alias, uri, data=delete_data, params=delete_params, headers=delete_headers,
                                         allow_redirects=delete_allow_redirects)
        if 'context' not in kwargs:
            self.delete_all_sessions()
        # 返回响应结果#
        return itf_result

    def __jsonPrettyOutput(self, json_str):
        '''
        预置条件：无
        函数功能：json串格式化输出
        参    数：json_str：原始json串
        返回值：格式化后的json串
        '''
        if isinstance(json_str, unicode) or isinstance(json_str, str):
            json_dict = self.jsonToDict(json_str)
        else:
            json_dict = dict(json_str)
        pretty_output = json.dumps(json_dict,
                                   ensure_ascii=False,
                                   sort_keys=True,
                                   indent=4,
                                   separators=(',', ':')
                                   )
        return pretty_output

    def messageFormatConvert(self, service, method, group, version, message_json):
        '''
        - 功能：将请求参数，转化为restful转化必须格式（中台专项）
        - 参数如下：
        - service：包名
        - method：方法名
        - group：组名
        - version：版本号
        - message_json：请求参数（列表类型）
        - RobotFramwork脚本写法示例如下：
        | ${restful转化参数} | messageFormatConvert | sendMmsMessage | MSG | 2.0.0 | 请求参数 |
        '''
        gao_message = {}
        gao_message["service"] = service
        gao_message["method"] = method
        gao_message["group"] = group
        gao_message["version"] = version
        gao_message["timeout"] = 3000
        gao_message["params"] = []
        # message_dict = self.jsonToDict(message_json)
        pretty_output = self.__jsonPrettyOutput(message_json)
        logger.info("##### Init parameter:\n" + pretty_output)
        gao_message["params"] = pretty_output
        # for item in message_dict:
        # elem_dict = {}
        # elem_dict["name"] = item
        # elem_dict["value"] = message_dict[item]
        # gao_message["params"].append(elem_dict)
        return json.dumps(gao_message)

    def jsonToDict(self, json_obj):
        '''
        - 功能：json串转化成字典类型
        - 参数如下：
        - json_obj：json字符串 
        - RobotFramwork脚本写法示例如下：
        | ${result_dict} | jsonToDict | ${response.content} |
        '''
        return json.loads(json_obj)

    def __checkDictKeyValueResult(self, item_count, typei="asc"):
        '''
        预置条件：已经调用__findDictKeyValue
        函数功能：根据__result_flag_itf，判断检查结果是否正确
        参    数：item_count：预期传入的key-value个数
        typei：asc-正向校验  desc-反向校验
        返回值：true/false，如果反向校验时，返回self.__result_flag_itf_desc
        '''
        sum_i = 0
        if self.__check_type == "asc":
            for item in self.__result_flag_itf.keys():
                sum_i = sum_i + self.__result_flag_itf[item]
        else:
            for item in self.__result_flag_itf_desc.keys():
                sum_i = sum_i + self.__result_flag_itf_desc[item]
        # 1.正向情况#
        if typei == "asc":
            # 1.1 如果找到的键值对数量，与预期需要找的键值对数量相等，则定为所有的预期键值对都找到#
            if item_count == sum_i:
                return True
            # 1.2 存在未找到预期key-value#
            else:
                return False
        # 2.反向情况#
        else:
            # 2.1 不希望找到，但是存在已找到的key-value#
            if sum > 0:
                return (False, self.__result_flag_itf_desc)
            # 2.2 符合预期，所有key-value都未出现#
            else:
                return (True, self.__result_flag_itf_desc)

    def __checkKeyInDict(self, dict_obj, key):
        '''
        预置条件：已经调用checkItfResult
        函数功能：判断字典中是否包含key
        参    数：dict_obj：字典
          key：要寻找的键
        返回值：无
        '''
        if isinstance(dict_obj, dict):
            for i in range(len(dict_obj)):
                # 取出当前遍历的key-value#
                key_now = dict_obj.keys()[i]
                value_now = dict_obj[key_now]
                # 如果当前的key与需要检查的key一致，说明找到键了#
                if key_now == key:
                    logger.info('##### Expect response contains key--->  ' + key + " check success. #####")
                    if self.__check_type == "asc":
                        self.__result_contain_flag.append(key)
                    else:
                        self.__result_contain_flag_desc.append(key)
                # 自我调用实现无限遍历#
                self.__checkKeyInDict(value_now, key)
        # 如果包含列表型#
        if isinstance(dict_obj, list):
            # 遍历列表中每个元素（字典）#
            for item in dict_obj:
                if item is not None:
                    # 把元素字典自调用#
                    self.__checkKeyInDict(item, key)

    def __findDictKeyValue(self, dict_obj, key, value, typei=__check_type):
        '''
        预置条件：已经调用接口
        函数功能：在json结果中，寻找key-value
        参    数：dict_obj：响应结果，经过json转dict
        key：要寻找的key
        value：要寻找的value
        typei：校验方式，默认是正向校验
        返回值：无
        '''
        # 使用isinstance检测数据类型为字典型#
        if isinstance(dict_obj, dict):
            for i in range(len(dict_obj)):
                # 取出当前遍历的key-value#
                key_now = dict_obj.keys()[i]
                value_now = dict_obj[key_now]
                # 如果当前的key与需要检查的key一致#
                if key_now == key:
                    # 如果当前的value与预期的value一致#
                    if isinstance(value_now, int) or isinstance(value_now, bool) or isinstance(value_now, float):
                        value_now = str(value_now)
                        # 布尔型比对时候忽略大小写#
                        if value_now.upper() == value.upper():
                            # 标记已经找到#
                            if typei == "asc":
                                logger.info(
                                    '##### Expect response contains key-value:%s ---> %s check success. #####' % (
                                        key, value))
                            # 将寻找到的结果，以key_value形式组成新的键，存放在结果标记变量中，防止重复#
                            tmp = key + "_" + value
                            if typei == "asc":
                                self.__result_flag_itf[tmp] = 1
                            else:
                                self.__result_flag_itf_desc[tmp] = 1
                            return
                    # 值为空列表时候，转换成字符串空[]#
                    if value_now == []:
                        value_now = '[]'
                    # 值为null(None)的时候，转化为字符串null#
                    if value_now is None:
                        value_now = 'null'
                    if value_now == value:
                        # 标记已经找到#
                        if typei == "asc":
                            logger.info('##### Expect response contains key-value:%s ---> %s check success. #####' % (
                                key, value))
                        # 将寻找到的结果，以key_value形式组成新的键，存放在结果标记变量中，防止重复#
                        tmp = key + "_" + value
                        if typei == "asc":
                            self.__result_flag_itf[tmp] = 1
                        else:
                            self.__result_flag_itf_desc[tmp] = 1
                        return
                    # 继续找#
                    else:
                        continue
                # 自我调用实现无限遍历#
                self.__findDictKeyValue(value_now, key, value, typei)
        # 如果包含列表型#
        if isinstance(dict_obj, list):
            # 遍历列表中每个元素（字典）#
            for item in dict_obj:
                if item is not None:
                    # 把元素字典自调用#
                    self.__findDictKeyValue(item, key, value, typei)

    def checkJsonResult(self, itf_result_json, *keyvalue):
        '''
        - 功能：WebSocket接口，返回结果仅为json格式情况下的校验
        - 参数如下：
        - itf_result_json：json格式响应结果
        - keyvalue：需要校验的返回的键值对
        - 校验符号说明：
        | =>   | 响应体中包含键值对       | name=>guozh    |
        - RobotFramwork脚本写法示例如下：
        | ${断言结果} | checkJsonResult | ${Json格式响应结果} | key1=>value1 | key2=>value2…… |
        '''
        # __result_flag_itf清空#
        self.__result_flag_itf = {}
        result_dict = self.jsonToDict(itf_result_json)
        for item in keyvalue:
            item_key_value = item.split("=>")
            # 调用键值对校验函数#
            self.__findDictKeyValue(result_dict,
                                    item_key_value[0],
                                    item_key_value[1]
                                    )
        if self.__checkDictKeyValueResult(len(keyvalue)):
            logger.info("##### Check Success. #####")
            return True
        else:
            tmp = []
            tmp_hope = {}
            # 遍历已经找到的键，放在临时tmp列表中#
            for item in self.__result_flag_itf:
                tmp.append(item.split("_")[0])
            # 遍历希望找到的键值，放在tmp_hope列表中#
            for hope_item in keyvalue:
                try:
                    tmp_hope[hope_item.split("=>")[0]] = hope_item.split("=>")[1]
                except:
                    pass
            for item in tmp_hope:
                # 第一种，如果预期的键都不在已经找到的键里面#
                if item not in tmp:
                    # 存在同名键的情况下，无此输出#
                    logger.info("##### [FAIL]Not find key-value:%s=>%s #####" % (item, tmp_hope[item]))
            return False

    def checkItfResult(self, itf_result, *keyvalue):
        '''
        - 功能：校验接口返回结果
        - 参数如下：
        - itf_result：调用HttpPost或者HttpGet方法，返回的响应结果
        - keyvalue：需要校验的返回的键值对
        - 校验符号说明：
        | 符号     |       说明                   |     示例                    |
        | *>>  | 校验http响应状态码     | *>>500         |
        | =>   | 响应体中包含键值对       | name=>guozh    |
        | =>>  | 响应头中包含键值对       | Server=>>xxs   |
        | (>   | 响应体中包含字符串       | (>woaini       |
        | (>>  | 响应头中包含字符串       | (>>keep-alive  |
        | ~>   | 响应体中包含某元素       | ~>Date         |
        | ~>>  | 响应头中包含某元素       | ~>>Connection  |
        | !=>  | 响应体中不包含键值对   | name!=>guozh    |
        | !=>> | 响应头中不包含键值对   | Server!=>>xxs   |
        | !(>  | 响应体中不包含字符串   | !(>woaini       |
        | !(>> | 响应头中不包含字符串   | !(>>keep-alive  |
        | !~>  | 响应体中不包含某元素   | !~>Date         |
        | !~>> | 响应头中不包含某元素   | !~>>Connection  |
        - RobotFramwork脚本写法示例如下：
        | ${断言结果} | checkItfResult | ${响应结果} | key1=>value1 | key2=>value2…… |
        '''
        # 转化成列表好修改#
        keyvalue = list(keyvalue)
        # __result_flag_itf清空#
        self.__result_flag_itf = {}
        # 校验包含的标志清空#
        self.__result_contain_flag = []
        self.tmp_find = {}
        itf_result_dict = {}
        # 记录包含校验信息#
        contain_flag_info = []
        keyvalue_asc_list = []  # 正向校验列表
        keyvalue_desc_list = []  # 反向校验列表
        # 默认期望的http响应码#
        hope_code = [200, 201, 204]
        # 键值对长度#
        pretty_output = ""
        pretty_output = self.__jsonPrettyOutput(itf_result.headers)
        logger.info("##### Response header:" + pretty_output)
        itf_result_header_str = json.dumps(dict(itf_result.headers)).replace("\", \"", "\",\"").replace("\": \"",
                                                                                                        "\":\"")
        if itf_result.content != "":
            # 响应结果正文转换成dict#
            try:
                itf_result_dict = self.jsonToDict(itf_result.content)
                # 响应结果格式化输出#
                pretty_output = self.__jsonPrettyOutput(itf_result.content)
                logger.info("##### Response message:\n" + pretty_output)
            except Exception, e:
                # 可能是服务器返回的错误堆栈信息#
                logger.warn(e)
                logger.warn("##### Response message is not json format. #####")
        else:
            logger.warn("##### Response don't contain json. #####")
        # 超级大校验（一）——校验响应码#
        # 列表赋值必须要用切片，否则是引用#
        tmp_list = keyvalue[:]
        code_err_list = []
        for item in tmp_list:
            if "*>>" in item:
                hopeCode = item.strip("*>>")
                if int(hopeCode) == itf_result.status_code:
                    logger.info("##### Http code check success ---> " + hopeCode + "find. #####")
                    # 将预期的响应代码追加到期望的响应码列表中，放开#
                    hope_code.append(int(hopeCode))
                else:
                    code_err_list.append(
                        "##### Http code check fail,expect-%s,actual-%s #####" % (hopeCode, itf_result.status_code))
                keyvalue.remove(item)
        # 超级大校验（二）#——包含不包含校验#
        tmp_list = keyvalue[:]
        err_list = []
        for item in tmp_list:
            # 响应结果体中的包含校验#
            if "(>" in item and "!(>" not in item and "!(>>" not in item and "(>>" not in item:
                hopeStr = item.strip("(>")
                if hopeStr in itf_result.content:
                    logger.info("##### Expect response contains content ---> " + hopeStr + " check success #####")
                else:
                    err_list.append("##### [FAIL]Response don't contain content ---> %s #####" % hopeStr)
                keyvalue.remove(item)
            if "!(>" in item and "!(>>" not in item:
                hopeStr = item.strip("!(>")
                if hopeStr in itf_result.content:
                    err_list.append(
                        "##### [FAIL]Expect response don't contain content ---> " + hopeStr + " actual exist. #####")
                logger.info("##### Expect response don't contain content ---> " + hopeStr + " check success. #####")
                keyvalue.remove(item)
            ##响应结果头中的包含校验#
            if "(>>" in item and "!(>>" not in item:
                hopeStr = item.strip("(>>")
                if hopeStr in itf_result_header_str:
                    logger.info(
                        "##### Expect response headers contains content ---> " + hopeStr + " check success. #####")
                else:
                    err_list.append("##### [FAIL]Response headers don't contain content ---> %s #####" % hopeStr)
                keyvalue.remove(item)
            if "!(>>" in item:
                hopeStr = item.strip("!(>>")
                if hopeStr in itf_result_header_str:
                    err_list.append(
                        "##### [FAIL]Expect response headers don't contain content ---> " + hopeStr + " but find. #####")
                logger.info(
                    "##### Expect response headers don't contain content ---> " + hopeStr + " check success. #####")
                keyvalue.remove(item)

        for item in keyvalue:
            if "!" in item:
                keyvalue_desc_list.append(item)
            else:
                keyvalue_asc_list.append(item)
        key_value_len = len(keyvalue_asc_list)
        key_value_len_desc = len(keyvalue_desc_list)
        # 兼容成都王珊珊的500场景，希望响应就是500情况（暂时保留）#
        if "\"status\": 500" in pretty_output:
            hope_code.append(500)
        # 超级大校验（三）具体键值对校验#
        desc_err_list = []
        if itf_result.status_code in hope_code:
            # 先校验反向情况#
            if key_value_len_desc > 0:
                item_key_value = []
                self.__check_type = "desc"
                for item in keyvalue_desc_list:
                    # 反向第一种校验方式：!=>校验方式#
                    if "!=>" in item and "!=>>" not in item:
                        item_key_value = item.split("!=>")
                        # 调用键值对校验函数#
                        self.__findDictKeyValue(itf_result_dict,
                                                item_key_value[0],
                                                item_key_value[1],
                                                "desc"
                                                )
                        flag = 0
                        for item in self.__result_flag_itf_desc:
                            tmp_list = item.split("_")
                            # fail情况#
                            if item_key_value[0] == tmp_list[0] and item_key_value[1] == tmp_list[1]:
                                flag = 1
                        if flag == 0:
                            logger.info(
                                "##### Expect response content don't contain key-value " + item_key_value[0] + "---> " +
                                item_key_value[1] + " check success. #####")
                    # 反向校验第二种校验方式：=>校验方式#
                    if "!=>>" in item:
                        item_key_value = item.split("!=>>")
                        # 调用键值对校验函数#
                        self.__findDictKeyValue(dict(itf_result.headers),
                                                item_key_value[0],
                                                item_key_value[1],
                                                "desc"
                                                )
                        flag = 0
                        for item in self.__result_flag_itf_desc:
                            tmp_list = item.split("_")
                            # fail情况#
                            if item_key_value[0] == tmp_list[0] and item_key_value[1] == tmp_list[1]:
                                flag = 1
                        if flag == 0:
                            logger.info(
                                "##### Expect response headers don't contain key-value " + item_key_value[0] + "---> " +
                                item_key_value[1] + " check success. #####")
                    # 反向校验第三种校验方式：！~>校验方式#
                    if "!~>" in item and "!~>>" not in item:
                        key = item.strip("!~>")
                        self.__checkKeyInDict(itf_result_dict, key)
                        # 报文中包含这个key#
                        if key in self.__result_contain_flag_desc:
                            desc_err_list.append(
                                "##### [FAIL]Expect response content don't contain key ---> " + key + " but find. #####")
                        else:
                            logger.info(
                                "##### Expect response content don't contain key ---> " + key + " check success. #####")
                    # 反向校验第四种校验方式：！~>>校验方式#
                    if "!~>>" in item:
                        key = item.strip("!~>>")
                        self.__checkKeyInDict(dict(itf_result.headers), key)
                        # 报文中包含这个key#
                        if key in self.__result_contain_flag_desc:
                            desc_err_list.append(
                                "##### [FAIL]Expect response content don't contain key ---> " + key + " but find. #####")
                        else:
                            logger.info(
                                "##### Expect respnse content don't contain key ---> " + key + " check success. #####")
                # 校验结果判断#
                result = self.__checkDictKeyValueResult(len(keyvalue), "desc")
                for item in result[1]:
                    tmp_list = item.split("_")
                    desc_err_list.append(
                        "##### [FAIL]Expect response content don't contain key-value " + tmp_list[0] + " ---> " +
                        tmp_list[1] + " but find. #####")
            if key_value_len > 0:
                # 还原校验方式为正向#
                self.__check_type = "asc"
                for item in keyvalue_asc_list:
                    # 第一种校验方式：=>校验方式#
                    if "=>" in item and "=>>" not in item:
                        item_key_value = item.split("=>")
                        # 调用键值对校验函数#
                        self.__findDictKeyValue(itf_result_dict,
                                                item_key_value[0],
                                                item_key_value[1]
                                                )
                    # 第二种校验方式：~>校验#
                    elif "~>" in item and "~>>" not in item:
                        key_value_len = key_value_len - 1
                        key = item.strip("~>")
                        self.__checkKeyInDict(itf_result_dict, key)
                        # 报文中包含这个key#
                        if key in self.__result_contain_flag:
                            pass
                        else:
                            contain_flag_info.append("##### [FAIL]Response don't contain key ---> %s #####" % key)
                    # 第三种校验方式：=>>校验#
                    elif "=>>" in item:
                        item_key_value = item.split("=>>")
                        # 调用键值对校验函数#
                        self.__findDictKeyValue(dict(itf_result.headers),
                                                item_key_value[0],
                                                item_key_value[1]
                                                )
                    # 第四种校验方式：~>>校验#
                    elif "~>>" in item:
                        key_value_len = key_value_len - 1
                        key = item.strip("~>>")
                        self.__checkKeyInDict(dict(itf_result.headers), key)
                        # 报文中包含这个key#
                        if key in self.__result_contain_flag:
                            pass
                        else:
                            contain_flag_info.append(
                                "##### [FAIL]Response headers don't contain key ---> %s #####" % key)
                    else:
                        logger.error("##### Unknow check symbol:%s #####" % item)
            # 校验结果判断#
            if self.__checkDictKeyValueResult(key_value_len):
                err_all = contain_flag_info + err_list + code_err_list + desc_err_list
                if err_all == []:
                    logger.info("##### Checkpoint successfully. #####")
                    return True
                else:
                    for item in err_all:
                        logger.info(item)
                    return False
            else:
                tmp = []
                tmp_hope = {}
                # 遍历已经找到的键，放在临时tmp列表中#
                for item in self.__result_flag_itf:
                    tmp.append(item.split("_")[0])
                # 遍历希望找到的键值，放在tmp_hope列表中#
                for hope_item in keyvalue_asc_list:
                    try:
                        tmp_hope[hope_item.split("=>")[0]] = hope_item.split("=>")[1]
                    except:
                        pass
                for item in tmp_hope:
                    # 第一种，如果预期的键都不在已经找到的键里面#
                    if item not in tmp:
                        # 注意！存在同名键的情况下，无此输出#
                        logger.info("##### [FAIL]Exist not found key-value:%s=>%s #####" % (item, tmp_hope[item]))
                err_all = contain_flag_info + err_list + code_err_list + desc_err_list
                if err_all != []:
                    for item in err_all:
                        logger.info(item)
                return False
            logger.info("##### Http code check success,but don't check response detail.")
            return True
        else:
            logger.info("##### Server response " + str(itf_result.status_code) + "error. #####")
            try:
                itf_result_dict = self.jsonToDict(itf_result.content)
                # 响应结果格式化输出#
                pretty_output = self.__jsonPrettyOutput(itf_result.content)
                logger.info(pretty_output)
            except Exception, e:
                logger.warn(e)
                logger.info(itf_result.content)
            return False


if __name__ == '__main__':
    # l = locust()
    #     l.createConfigure("http://10.0.0.205:8088",
    #                       "transform",
    #                       data='{ "group":"user-center-service","method":"authUserByMobile","params":["3UDO5RAHPMBQEODP7ILJV3LLZY3ZK","18700000006","123456"],"service":"cn.com.flaginfo.platform.ucenter.api.AuthService","timeout":3000, "version":"6.1.0"}',
    #                       no_web="False",
    #                       checkpoint="message=>用户验证成功！"
    #                       )
    # l.createConfigure("http://www.sojson.com",
    #                   "/open/api/weather/json.shtml?city=南京",
    #                   type="GET",
    #                   header='{"Content-type":"application/x-www-form-urlencoded;charset=UTF-8"}')
    # l.execLocust()

    upload_itf = http()
    login_result= upload_itf.HttpPost(url='http://ep-portal.test5.ums86.com', uri='/user/login.do',headers='{"Content-Type":"application/x-www-form-urlencoded;charset=UTF-8"}',data="spCode=flmFp6upCYReMsJuKVGfTA%3D%3D&userName=yfDXU7BnWU70ZutL4r7FZQ%3D%3D&password=461aoAmPRNhrCpHJ%2BOuSyg%3D%3D&randCode=2g0SqLcNBml6CS7gXHCjLA%3D%3D&_scrtk=1566899664153705&keepPwd=uBmhuBrwrojVKp8Iit3clg%3D%3D")
    login_cookie = login_result.headers['Set-Cookie']

    # print login_cookie
    # upload_result = upload_itf.HttpPost(url='http://ep-portal.test5.ums86.com', uri='/common/dynaData!upload.do',
    #                                     headers='{"Content-Type":"multipart/form-data","Cookie":"%s"}' % login_cookie,
    #                                     files='{"dynaDataFile":"/Users/xiaoliwei/Downloads/aa.csv"}')
    # print upload_result.content

    import requests
    url = "http://ep-portal.test5.ums86.com/common/dynaData!upload.do"
    header = {'Cookie':'%s'%login_cookie}
    print header
    files ={'dynaDataFile':('aa.csv',open('/Users/xiaoliwei/Downloads/aa.csv','rb'),'multipart/form-data')}
    response = requests.post(url, files=files, headers=header)
    print response.status_code, response.text





